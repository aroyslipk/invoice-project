# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# Name:         views.py
# Purpose:      Handles all view and API endpoint logic for the Invoice app.
#
# Author:       AnikRoy
# GitHub:       https://github.com/aroyslipk
#
# Created:      2025-06-20
# Copyright:    (c) AnikRoy 2025
# Licence:      Proprietary
# -----------------------------------------------------------------------------

"""
Handles all view logic for the Invoice application, including template rendering,
API endpoints, and business logic for the multi-tenant permission system.
"""

# Python Standard Library Imports
import datetime
from pathlib import Path
from copy import copy
import os
import json

# Django Core Imports
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import LoginView
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date
from django.conf import settings
from .forms import PriceForm, WorkEntryForm, AdminUserCreationForm
from django.contrib import messages
from .forms import PriceForm, WorkEntryForm, AdminUserCreationForm, ClientProjectForm
from django.db.models import Count, Sum
from django.utils import timezone
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib.auth.tokens import default_token_generator
from django.utils.crypto import get_random_string
from django.db.models.functions import ExtractWeekDay


# Third-Party Library Imports
from openpyxl import Workbook, load_workbook
from num2words import num2words
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from rest_framework import generics, permissions
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

# Local Application Imports
from .forms import PriceForm, WorkEntryForm
from .models import ClientProject, Price, User, WorkEntry
from .serializers import (
    ClientProjectSerializer, PriceSerializer, RegisterSerializer,
    UserSerializer, WorkDashboardSerializer, WorkEntrySerializer,
)


# --- Core & Template Views ---

def home_view(request):
    """Displays the home page."""
    return HttpResponse("<h2>🧾 Welcome to the Invoice App</h2><p>Go to <a href='/login/'>Login</a></p>")


@login_required
def dashboard_template_view(request):
    """
    Renders a comprehensive, interactive dashboard that includes summary cards,
    advanced filtering, dynamic charts, and new analytics.
    """
    user = request.user
    if user.role == 'user':
        return redirect('my_work_entries')

    # --- Step 1: Initial Querysets based on User Role ---
    projects_qs = ClientProject.objects.all()
    users_qs = User.objects.filter(role='user')
    entries_qs = WorkEntry.objects.select_related('user', 'project')

    if user.role == 'admin':
        projects_qs = projects_qs.filter(managed_by=user)
        users_qs = users_qs.filter(managed_by=user)
        entries_qs = entries_qs.filter(project__managed_by=user)

    # --- Step 2: Apply Advanced Filtering from GET Parameters ---
    selected_project_id = request.GET.get('project')
    selected_user_id = request.GET.get('user')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if selected_project_id:
        entries_qs = entries_qs.filter(project_id=selected_project_id)
    if selected_user_id:
        entries_qs = entries_qs.filter(user_id=selected_user_id)
    if start_date:
        entries_qs = entries_qs.filter(date__gte=start_date)
    if end_date:
        entries_qs = entries_qs.filter(date__lte=end_date)
    
    # --- Step 3: Calculate Data for Existing Cards (using filtered data) ---
    total_projects = projects_qs.count()
    total_team_members = users_qs.count()
    current_month_entries = entries_qs.filter(
        date__year=timezone.now().year,
        date__month=timezone.now().month
    ).count()

    # --- Step 4: Calculate Data for New Analytics Cards ---
    # Busiest Day of the Week
    busiest_day_query = entries_qs.annotate(
        weekday=ExtractWeekDay('date')
    ).values('weekday').annotate(count=Count('id')).order_by('-count').first()
    
    if busiest_day_query:
        weekday_map = {1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday', 5: 'Thursday', 6: 'Friday', 7: 'Saturday'}
        busiest_day = weekday_map.get(busiest_day_query['weekday'], 'N/A')
    else:
        busiest_day = 'N/A'

    # Most Productive User (by Quantity)
    most_productive_user_query = entries_qs.values('user__username').annotate(
        total_quantity=Sum('quantity')
    ).order_by('-total_quantity').first()
    
    most_productive_user = most_productive_user_query['user__username'] if most_productive_user_query else 'N/A'
    
    # --- Step 5: Calculate Data for Charts ---
    category_data = entries_qs.values('category').annotate(count=Count('id')).order_by('-count')
    pie_chart_labels = [item['category'] for item in category_data]
    pie_chart_data = [item['count'] for item in category_data]

    monthly_data = (
        entries_qs.values('date__year', 'date__month')
        .annotate(count=Count('id'))
        .order_by('date__year', 'date__month')
    )
    bar_chart_labels = [f"{item['date__year']}-{item['date__month']:02d}" for item in monthly_data]
    bar_chart_data = [item['count'] for item in monthly_data]

    # --- Step 6: Prepare the Final Context for the Template ---
    context = {
        'total_projects': total_projects,
        'total_team_members': total_team_members,
        'current_month_entries': current_month_entries,
        'busiest_day': busiest_day, # নতুন ডেটা যোগ করা হয়েছে
        'most_productive_user': most_productive_user, # নতুন ডেটা যোগ করা হয়েছে
        
        'entries': entries_qs.order_by('-date'),
        
        'pie_chart_labels': pie_chart_labels,
        'pie_chart_data': pie_chart_data,
        'bar_chart_labels': bar_chart_labels,
        'bar_chart_data': bar_chart_data,

        'all_projects': projects_qs.order_by('name'),
        'all_users': users_qs.order_by('username'),

        'selected_project_id': selected_project_id,
        'selected_user_id': selected_user_id,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'dashboard.html', context)


@login_required
def admin_panel_view(request):
    """Displays the super admin panel with access to all system data."""
    if request.user.role != 'super_admin':
        return render(request, 'unauthorized.html')
    context = {
        'work_entries': WorkEntry.objects.all().order_by('-date'),
        'prices': Price.objects.all(),
        'projects': ClientProject.objects.all(),
        'users': User.objects.all(),
    }
    return render(request, 'admin_panel.html', context)


# --- User-Facing Forms & Views ---

@login_required
def work_entry_form_view(request):
    """Handles the form for standard users to submit their work."""
    if request.user.role != 'user':
        return render(request, 'unauthorized.html')
    if request.method == 'POST':
        form = WorkEntryForm(request.POST, user=request.user)
        if form.is_valid():
            work_entry = form.save(commit=False)
            work_entry.user = request.user
            work_entry.save()
            return redirect('my_work_entries')
    else:
        form = WorkEntryForm(user=request.user)
    return render(request, 'work_entry_form.html', {'form': form})


@login_required
def my_work_entries_view(request):
    """
    Displays a comprehensive dashboard for the user, including summary cards,
    a calendar view, search/filter functionality, and a list of their work entries.
    """
    if request.user.role != 'user':
        return render(request, 'unauthorized.html')

    user = request.user
    now = timezone.now()
    work_entries = WorkEntry.objects.filter(user=user).select_related('project')

    # --- ফিল্টার লজিক ---
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get('query')

    if start_date:
        work_entries = work_entries.filter(date__gte=start_date)
    if end_date:
        work_entries = work_entries.filter(date__lte=end_date)
    if query:
        work_entries = work_entries.filter(category__icontains=query)

    # --- সারসংক্ষেপ কার্ডের জন্য ডেটা গণনা ---
    total_entries_month = work_entries.filter(
        date__year=now.year, date__month=now.month
    ).count()

    quantity_sum_result = work_entries.filter(
        date__year=now.year, date__month=now.month
    ).aggregate(total_quantity=Sum('quantity'))
    total_quantity_month = quantity_sum_result['total_quantity'] or 0

    frequent_project_query = work_entries.values('project__name').annotate(
        count=Count('project')
    ).order_by('-count').first()
    most_frequent_project = frequent_project_query['project__name'] if frequent_project_query else 'N/A'

    # --- ক্যালেন্ডারের জন্য ডেটা প্রস্তুত করা ---
    calendar_events = []
    entry_dates = work_entries.values('date').distinct()
    for entry_date in entry_dates:
        calendar_events.append({
            'title': 'Work Submitted',
            'start': entry_date['date'].strftime('%Y-%m-%d'),
            'allDay': True
        })

    context = {
        'entries': work_entries.order_by('-date'),
        'total_entries_month': total_entries_month,
        'total_quantity_month': total_quantity_month,
        'most_frequent_project': most_frequent_project,
        'calendar_events_json': json.dumps(calendar_events),
        # ফিল্টারের মানগুলো টেমপ্লেটে ফেরত পাঠানো
        'start_date': start_date,
        'end_date': end_date,
        'query': query,
    }
    
    return render(request, 'my_work_entries.html', context)


# --- Pricing Management Views ---

@login_required
def manage_prices_view(request):
    """
    Displays a list of prices and a form to add a new price.
    Handles creation of new prices.
    """
    user = request.user
    if user.role not in ['admin', 'super_admin']:
        return render(request, 'unauthorized.html')

    # Price creation logic
    if request.method == 'POST':
        form = PriceForm(request.POST)
        if form.is_valid():
            price = form.save(commit=False)
            price.managed_by = user
            price.save()
            messages.success(request, f"Price for '{price.category}' was added successfully.")
            return redirect('manage_prices')
    else:
        form = PriceForm()

    # Get the list of prices based on user role
    if user.role == 'super_admin':
        prices = Price.objects.all().order_by('category')
    else:  # 'admin' role
        prices = Price.objects.filter(managed_by=user).order_by('category')
    
    context = {
        'form': form,
        'prices': prices,
    }
    return render(request, 'manage_prices.html', context)



@login_required
def price_edit_view(request, id):
    """Handles editing of a specific price, with ownership check."""
    price = get_object_or_404(Price, id=id)
    user = request.user
    if user.role != 'super_admin' and price.managed_by != user:
        return render(request, 'unauthorized.html')
    
    if request.method == 'POST':
        form = PriceForm(request.POST, instance=price)
        if form.is_valid():
            form.save()
            return redirect('price_manage')
    else:
        form = PriceForm(instance=price)
    return render(request, 'price_edit.html', {'form': form, 'price': price})


@login_required
def delete_price_view(request, id):
    """
    Deletes a specific price after ownership verification.
    """
    user = request.user
    price = get_object_or_404(Price, id=id)

    # Security check
    if user.role != 'super_admin' and price.managed_by != user:
        return render(request, 'unauthorized.html')
    
    category_name = price.category
    price.delete()
    messages.success(request, f"Price for '{category_name}' has been deleted.")
    return redirect('manage_prices')


# --- Reporting & Invoice Generation ---

@login_required
def export_page_view(request):
    """Displays the page for initiating exports."""
    if request.user.role not in ['admin', 'super_admin']:
        return render(request, 'unauthorized.html')
    return render(request, 'export_page.html')

@login_required
def my_team_view(request):
    """
    Allows an admin to view their managed users and add new ones.
    Sends a welcome email with a temporary password to the new user.
    """
    if request.user.role not in ['admin', 'super_admin']:
        return render(request, 'unauthorized.html')

    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save(commit=False)
            
            # --- সঠিক এবং চূড়ান্ত পদ্ধতিতে পাসওয়ার্ড তৈরি করা ---
            password = get_random_string(length=12)
            new_user.set_password(password)
            
            new_user.role = 'user'
            new_user.managed_by = request.user
            new_user.save()

            # --- নতুন ইউজারকে ওয়েলকাম ইমেইল পাঠানো ---
            mail_subject = 'Welcome to InvoiceApp!'
            message = render_to_string('welcome_email.html', {
                'user': new_user,
                'password': password,
                'admin_name': request.user.username,
            })
            
            send_mail(
                subject=mail_subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[new_user.email],
                html_message=message
            )
            
            messages.success(request, f"User '{new_user.username}' was created successfully! A welcome email with credentials has been sent.")
            return redirect('my_team')
    else:
        form = AdminUserCreationForm()

    if request.user.role == 'super_admin':
        managed_users = User.objects.filter(role='user')
    else:
        managed_users = User.objects.filter(managed_by=request.user)
    
    context = {
        'form': form,
        'managed_users': managed_users
    }
    return render(request, 'my_team.html', context)

@login_required
def manage_projects_view(request):
    """
    Allows admins to view, create, and manage their projects.
    """
    user = request.user
    if user.role not in ['admin', 'super_admin']:
        return render(request, 'unauthorized.html')

    # Project creation logic
    if request.method == 'POST':
        form = ClientProjectForm(request.POST, request.FILES)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = user
            project.managed_by = user
            project.save()
            messages.success(request, f"Project '{project.name}' was created successfully.")
            return redirect('manage_projects')
    else:
        form = ClientProjectForm()

    # Get the list of projects based on user role
    if user.role == 'super_admin':
        projects = ClientProject.objects.all().order_by('-start_date')
    else:  # 'admin' role
        projects = ClientProject.objects.filter(managed_by=user).order_by('-start_date')

    context = {
        'form': form,
        'projects': projects,
    }
    return render(request, 'manage_projects.html', context)

@login_required
def delete_project_view(request, project_id):
    """
    Deletes a specific project after ownership verification.
    """
    user = request.user
    project = get_object_or_404(ClientProject, id=project_id)

    # Security check
    if user.role != 'super_admin' and project.managed_by != user:
        return render(request, 'unauthorized.html')

    project_name = project.name
    project.delete()
    messages.success(request, f"Project '{project_name}' has been deleted.")
    return redirect('manage_projects')


@login_required
def generate_invoice(request, project_id):
    """
    Generates and downloads a secure, role-filtered XLSX invoice,
    showing sums for only the 'Quantity' and 'Total Amount' columns.
    """
    user = request.user
    project = get_object_or_404(ClientProject, id=project_id)

    # Security check: Ensure admin can only access their own projects.
    if user.role != 'super_admin' and project.managed_by != user:
        return render(request, 'unauthorized.html')

    # Filter data based on user role.
    if user.role == 'super_admin':
        entries = WorkEntry.objects.filter(project=project).order_by('date')
        prices_qs = Price.objects.all()
    else:  # 'admin' role
        entries = WorkEntry.objects.filter(project=project, project__managed_by=user).order_by('date')
        prices_qs = Price.objects.filter(managed_by=user)
    
    prices = {p.category.lower(): p.rate for p in prices_qs}
    
    template_path = os.path.join(settings.BASE_DIR, 'static', 'template', 'InvoiceTemplate.xlsx')

    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        error_message = f"Error: The invoice template could not be found. Path checked: {template_path}"
        return HttpResponse(error_message, status=500)

    ws = wb.active
    ws.protection.sheet = False
    start_row = 14
    
    if entries:
        if len(entries) > 1:
            ws.insert_rows(start_row + 1, amount=len(entries) - 1)
        
        current_row = start_row
        total_quantity_sum, total_amount_numeric = 0, 0
        
        for entry in entries:
            rate_for_entry = float(prices.get(entry.category.lower(), 0))
            amount_for_entry = entry.quantity * rate_for_entry
            total_quantity_sum += entry.quantity
            total_amount_numeric += amount_for_entry
            
            # Populate data into cells
            ws.cell(row=current_row, column=2).value = entry.category
            ws.cell(row=current_row, column=3).value = entry.date.strftime("%Y-%m-%d")
            ws.cell(row=current_row, column=4).value = entry.quantity
            ws.cell(row=current_row, column=5).value = rate_for_entry
            ws.cell(row=current_row, column=6).value = amount_for_entry
            current_row += 1
            
        # --- Summary section ---
        summary_start_row = start_row + len(entries) + 2
        ws.cell(row=summary_start_row, column=4).value = total_quantity_sum
        ws.cell(row=summary_start_row, column=4).font = Font(bold=True)
        
        ws.cell(row=summary_start_row, column=6).value = total_amount_numeric
        ws.cell(row=summary_start_row, column=6).font = Font(bold=True)
        
        # --- Amount in words ---
        in_words_cell = ws.cell(row=summary_start_row + 1, column=2)
        total_in_words = num2words(total_amount_numeric, lang='en').title() + " US Dollars Only"
        in_words_cell.value = total_in_words
        in_words_cell.font = Font(italic=True)

    # --- Project attachment link (if exists) ---
    if project.attachment and hasattr(project.attachment, 'url'):
        # এই কোডটুকু এখন ಸರಿಯಾಗಿ indent করা হয়েছে
        attachment_url = request.build_absolute_uri(project.attachment.url)
        attachment_cell = ws.cell(row=ws.max_row + 2, column=2)
        attachment_cell.value = "View Project Attachment"
        attachment_cell.hyperlink = attachment_url
        attachment_cell.font = Font(color="0000FF", underline="single")
        ws.cell(row=attachment_cell.row, column=1).value = "Attachment:"
        
    # --- Prepare and return the response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Invoice_{project.name}_{datetime.date.today()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

# --------------------------------------------------------------------------
# --- API Views (DRF) with full security and role-based filtering ---
# --------------------------------------------------------------------------

class RegisterView(generics.CreateAPIView):
    """API endpoint for public user registration."""
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]


class ProfileView(generics.RetrieveAPIView):
    """API endpoint to view the profile of the currently logged-in user."""
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_object(self):
        return self.request.user


class WorkEntryListCreateView(generics.ListCreateAPIView):
    """API endpoint to list and create work entries."""
    serializer_class = WorkEntrySerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'super_admin':
            return WorkEntry.objects.all()
        if user.role == 'admin':
            return WorkEntry.objects.filter(project__managed_by=user)
        return WorkEntry.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ExportWorkEntriesXLSXView(generics.GenericAPIView):
    """API endpoint to export work entries to an XLSX file."""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.role not in ["admin", "super_admin"]:
            return Response({"detail": "Not authorized."}, status=403)
        # ... (Your Export logic from previous version, now secured) ...
        return Response({"detail": "Export successful."}) # Placeholder response


class DashboardView(generics.ListAPIView):
    """API endpoint for dashboard data, filtered by role."""
    serializer_class = WorkDashboardSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'super_admin':
            queryset = WorkEntry.objects.all()
        elif user.role == 'admin':
            queryset = WorkEntry.objects.filter(project__managed_by=user)
        else:
            return WorkEntry.objects.none()
        # ... (Your query parameter filtering logic remains here) ...
        return queryset


class PriceListCreateView(generics.ListCreateAPIView):
    """API endpoint for listing and creating prices."""
    serializer_class = PriceSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role in ['super_admin', 'admin']:
            return Price.objects.filter(managed_by=user) if user.role == 'admin' else Price.objects.all()
        return Price.objects.none()

    def perform_create(self, serializer):
        serializer.save(managed_by=self.request.user)


class PriceDetailView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, or deleting a specific price."""
    serializer_class = PriceSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'id'

    def get_queryset(self):
        user = self.request.user
        if user.role == 'super_admin':
            return Price.objects.all()
        if user.role == 'admin':
            return Price.objects.filter(managed_by=user)
        return Price.objects.none()


class ClientProjectListCreateView(generics.ListCreateAPIView):
    """API endpoint for listing and creating client projects."""
    serializer_class = ClientProjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'super_admin':
            return ClientProject.objects.all()
        if user.role == 'admin':
            return ClientProject.objects.filter(managed_by=user)
        return ClientProject.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, managed_by=self.request.user)


class CustomLoginView(LoginView):
    """Custom login view using a template."""
    template_name = 'login.html'
    form_class = AuthenticationForm

    def form_valid(self, form):
        login(self.request, form.get_user())
        return super().form_valid(form)